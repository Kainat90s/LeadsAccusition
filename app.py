~"""
Production-Ready Lead Scraping System
=====================================
Uses Google APIs for authentic business discovery and HTML parsing for contact extraction.
Designed for software house lead generation and email marketing.
"""

import os
import re
import time
import logging
import concurrent.futures
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Set
from urllib.parse import urlparse, urljoin

import dns.resolver
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

# Try to import Playwright for browser automation (fallback for bot-protected sites)
PLAYWRIGHT_AVAILABLE = False
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    logging.warning("Playwright not installed. Install with: pip install playwright && playwright install chromium")
    pass

# --- CONFIGURATION ---
app = Flask(__name__)
CORS(app)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# --- CONSTANTS ---
MASTER_DB_FILE = 'Scraping.xlsx'
SHEET_NAME = 'Leads'

# Required output field names (DO NOT MODIFY)
EXCEL_HEADERS = [
    'Company Name',
    'Company website',
    'Company Email',
    'Email status',
    'Company Phone Number',
    'Linkedin Link',
    'X/ Twitter Link',
    'Facebook Link',
    'Instagram Link',
    'Tiktok Link',
    'YouTube channel Link',
]

# --- API KEYS ---
GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")
GOOGLE_CSE_ID = os.environ.get("GOOGLE_CSE_ID", "")  # Custom Search Engine ID
GOOGLE_PLACES_API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY", "") or GOOGLE_API_KEY

# --- REQUEST CONFIGURATION ---
REQUEST_TIMEOUT = 15
MAX_RETRIES = 3
RETRY_DELAY = 1

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
]

def get_headers():
    """Get request headers with rotating user agent."""
    import random
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "keep-alive",
    }


# --- REGEX PATTERNS ---
# Improved email pattern - catches more formats including obfuscated emails
EMAIL_PATTERN = re.compile(
    r'[A-Za-z0-9._%+-]+\s*[@\[at\]]\s*[A-Za-z0-9.-]+\s*[.\[dot\]]\s*[A-Za-z]{2,}',
    re.I
)

# Standard email pattern for clean extraction
EMAIL_PATTERN_STRICT = re.compile(
    r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b'
)

# Phone patterns - comprehensive for international formats
PHONE_PATTERNS = [
    # US/Canada: (123) 456-7890, 123-456-7890, 123.456.7890
    re.compile(r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}'),
    # International with country code: +1 123 456 7890, +44 20 1234 5678
    re.compile(r'\+\d{1,3}[-.\s]?\(?\d{1,4}\)?[-.\s]?\d{1,4}[-.\s]?\d{1,4}[-.\s]?\d{0,4}'),
    # UK format: 020 1234 5678, 07123 456789
    re.compile(r'0\d{2,4}[-.\s]?\d{3,4}[-.\s]?\d{3,4}'),
    # Generic international: various formats
    re.compile(r'\+?\d{1,4}[-.\s]?\d{2,4}[-.\s]?\d{2,4}[-.\s]?\d{2,4}'),
    # Toll-free: 1-800-123-4567
    re.compile(r'1[-.\s]?8[0-9]{2}[-.\s]?\d{3}[-.\s]?\d{4}'),
]

SOCIAL_PATTERNS = {
    'linkedin': re.compile(r'https?://(?:www\.)?linkedin\.com/(?:company|in)/[^\s\'"<>]+', re.I),
    'twitter': re.compile(r'https?://(?:www\.)?(?:twitter\.com|x\.com)/[^\s\'"<>]+', re.I),
    'facebook': re.compile(r'https?://(?:www\.)?(?:facebook|fb)\.com/[^\s\'"<>]+', re.I),
    'instagram': re.compile(r'https?://(?:www\.)?instagram\.com/[^\s\'"<>]+', re.I),
    'tiktok': re.compile(r'https?://(?:www\.)?tiktok\.com/@[^\s\'"<>]+', re.I),
    'youtube': re.compile(r'https?://(?:www\.)?(?:youtube\.com/(?:channel|c|user|@)|youtu\.be/)[^\s\'"<>]+', re.I),
}

# Contact page URL patterns - expanded
CONTACT_PAGE_PATTERNS = [
    '/contact', '/contact-us', '/contactus', '/contact_us', '/contact.html', '/contact.php',
    '/get-in-touch', '/reach-us', '/connect', '/reach-out',
    '/support', '/help', '/customer-service', '/customer-support',
    '/enquiry', '/inquiry', '/enquiries', '/inquiries',
    '/feedback', '/write-to-us', '/talk-to-us',
    '/locations', '/location', '/offices', '/office',
]

ABOUT_PAGE_PATTERNS = [
    '/about', '/about-us', '/aboutus', '/about_us', '/about.html', '/about.php',
    '/who-we-are', '/our-story', '/company', '/our-company',
    '/team', '/our-team', '/leadership', '/management',
    '/profile', '/company-profile', '/corporate',
]

# Additional pages that often contain contact info
ADDITIONAL_CONTACT_PAGES = [
    '/footer', '/sitemap', '/privacy', '/privacy-policy',
    '/terms', '/legal', '/imprint', '/impressum',
]


@dataclass
class CompanyLead:
    """Data class for a single company lead."""
    company_name: str = ""
    company_website: str = ""
    company_email: str = ""
    email_status: str = "Not verified"
    company_phone: str = ""
    linkedin_link: str = ""
    twitter_link: str = ""
    facebook_link: str = ""
    instagram_link: str = ""
    tiktok_link: str = ""
    youtube_link: str = ""
    all_emails: List[str] = field(default_factory=list)
    all_phones: List[str] = field(default_factory=list)

    def to_excel_row(self) -> List[str]:
        """Convert to Excel row with exact field names."""
        return [
            self.company_name,
            self.company_website,
            self.company_email,
            self.email_status,
            self.company_phone,
            self.linkedin_link,
            self.twitter_link,
            self.facebook_link,
            self.instagram_link,
            self.tiktok_link,
            self.youtube_link,
        ]

    def to_dict(self) -> Dict:
        """Convert to dictionary for API response."""
        return {
            'Company Name': self.company_name,
            'Company website': self.company_website,
            'Company Email': self.company_email,
            'Email status': self.email_status,
            'Company Phone Number': self.company_phone,
            'Linkedin Link': self.linkedin_link,
            'X/ Twitter Link': self.twitter_link,
            'Facebook Link': self.facebook_link,
            'Instagram Link': self.instagram_link,
            'Tiktok Link': self.tiktok_link,
            'YouTube channel Link': self.youtube_link,
        }


# =============================================================================
# GOOGLE API FUNCTIONS - Primary Business Discovery
# =============================================================================

def search_companies_google(query: str, limit: int = 10) -> List[Dict]:
    """
    Primary function to discover companies using Google APIs.
    Uses Google Custom Search API and Google Places API.
    Returns list of companies with name and website URL.
    """
    companies = []
    seen_domains: Set[str] = set()

    # Try Google Custom Search API first
    cse_results = _search_google_custom_search(query, limit)
    for company in cse_results:
        domain = _extract_domain(company.get('website', ''))
        if domain and domain not in seen_domains:
            seen_domains.add(domain)
            companies.append(company)

    # Supplement with Google Places API
    if len(companies) < limit:
        places_results = _search_google_places(query, limit - len(companies))
        for company in places_results:
            domain = _extract_domain(company.get('website', ''))
            if domain and domain not in seen_domains:
                seen_domains.add(domain)
                companies.append(company)

    return companies[:limit]


def _search_google_custom_search(query: str, limit: int) -> List[Dict]:
    """
    Search using Google Custom Search JSON API.
    Requires GOOGLE_API_KEY and GOOGLE_CSE_ID.
    """
    results = []
    if not GOOGLE_API_KEY or not GOOGLE_CSE_ID:
        logging.warning("Google Custom Search API credentials not configured")
        return results

    try:
        url = "https://www.googleapis.com/customsearch/v1"
        params = {
            'key': GOOGLE_API_KEY,
            'cx': GOOGLE_CSE_ID,
            'q': query,
            'num': min(limit, 10),  # API max is 10 per request
        }

        response = _make_request_with_retry(url, params=params, method='GET')
        if response and response.status_code == 200:
            data = response.json()
            for item in data.get('items', []):
                link = item.get('link', '')
                if link and _is_valid_company_website(link):
                    results.append({
                        'name': item.get('title', '').split(' - ')[0].split(' | ')[0].strip(),
                        'website': link,
                        'source': 'Google Custom Search'
                    })
    except Exception as e:
        logging.error(f"Google Custom Search error: {e}")

    return results


def _search_google_places(query: str, limit: int) -> List[Dict]:
    """
    Search using Google Places API (Text Search).
    Returns companies with websites.
    """
    results = []
    api_key = GOOGLE_PLACES_API_KEY or GOOGLE_API_KEY
    if not api_key:
        logging.warning("Google Places API key not configured")
        return results

    try:
        # Text Search API
        url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
        params = {
            'query': query,
            'key': api_key,
        }

        response = _make_request_with_retry(url, params=params, method='GET')
        if response and response.status_code == 200:
            data = response.json()
            for place in data.get('results', [])[:limit]:
                place_id = place.get('place_id')
                if place_id:
                    details = _get_place_details(place_id, api_key)
                    if details and details.get('website'):
                        results.append({
                            'name': details.get('name', place.get('name', '')),
                            'website': details.get('website'),
                            'phone': details.get('formatted_phone_number', ''),
                            'source': 'Google Places'
                        })
    except Exception as e:
        logging.error(f"Google Places search error: {e}")

    return results


def _get_place_details(place_id: str, api_key: str) -> Optional[Dict]:
    """Get detailed information about a place including website."""
    try:
        url = "https://maps.googleapis.com/maps/api/place/details/json"
        params = {
            'place_id': place_id,
            'fields': 'name,website,formatted_phone_number,formatted_address',
            'key': api_key,
        }

        response = _make_request_with_retry(url, params=params, method='GET')
        if response and response.status_code == 200:
            data = response.json()
            return data.get('result', {})
    except Exception as e:
        logging.error(f"Place details error: {e}")
    return None


def get_company_websites(companies: List[Dict]) -> List[Dict]:
    """
    Filter and validate company websites from search results.
    Returns only companies with valid, accessible websites.
    """
    valid_companies = []

    for company in companies:
        website = company.get('website', '')
        if website and _is_valid_company_website(website):
            # Normalize URL
            if not website.startswith(('http://', 'https://')):
                website = 'https://' + website
            company['website'] = website
            valid_companies.append(company)

    return valid_companies


# =============================================================================
# WEBSITE DATA EXTRACTION - HTML Parsing
# =============================================================================

def scrape_website_contact_info(company: Dict, use_browser_fallback: bool = True) -> CompanyLead:
    """
    Main function to scrape contact information from a company website.
    Comprehensive extraction from homepage, contact page, about page, and more.
    Uses browser automation as fallback for bot-protected sites.
    """
    lead = CompanyLead(
        company_name=company.get('name', ''),
        company_website=company.get('website', ''),
    )

    website = company.get('website', '')
    if not website:
        return lead

    # Pre-populated phone from Google Places
    if company.get('phone'):
        cleaned_phone = _clean_phone_number(company['phone'])
        if cleaned_phone:
            lead.all_phones.append(cleaned_phone)

    visited_urls = set()
    homepage_html = None
    used_browser = False

    try:
        # =====================================================================
        # PHASE 1: Try with regular HTTP requests (fast)
        # =====================================================================
        
        # 1. Fetch and parse homepage
        logging.info(f"Scraping homepage: {website}")
        homepage_html = _fetch_page(website, use_browser=False)
        if homepage_html:
            _extract_from_html(homepage_html, website, lead)
            visited_urls.add(website)

        # 2. Fetch and parse contact page (try multiple patterns)
        contact_url = _find_contact_page(website, homepage_html)
        if contact_url and contact_url not in visited_urls:
            logging.info(f"Scraping contact page: {contact_url}")
            contact_html = _fetch_page(contact_url, use_browser=False)
            if contact_html:
                _extract_from_html(contact_html, contact_url, lead)
                visited_urls.add(contact_url)

        # 3. Fetch and parse about page
        about_url = _find_about_page(website, homepage_html)
        if about_url and about_url not in visited_urls:
            logging.info(f"Scraping about page: {about_url}")
            about_html = _fetch_page(about_url, use_browser=False)
            if about_html:
                _extract_from_html(about_html, about_url, lead)
                visited_urls.add(about_url)

        # 4. If still no email/phone, try additional pages
        if not lead.all_emails or not lead.all_phones:
            additional_pages = _find_additional_contact_pages(website, homepage_html)
            for page_url in additional_pages[:3]:
                if page_url not in visited_urls:
                    logging.info(f"Scraping additional page: {page_url}")
                    page_html = _fetch_page(page_url, use_browser=False)
                    if page_html:
                        _extract_from_html(page_html, page_url, lead)
                        visited_urls.add(page_url)
                    if lead.all_emails and lead.all_phones:
                        break

        # 5. Try common contact page URLs directly
        if not lead.all_emails:
            for pattern in CONTACT_PAGE_PATTERNS[:5]:
                test_url = urljoin(website, pattern)
                if test_url not in visited_urls:
                    page_html = _fetch_page(test_url, use_browser=False)
                    if page_html:
                        _extract_from_html(page_html, test_url, lead)
                        visited_urls.add(test_url)
                        if lead.all_emails:
                            break

        # =====================================================================
        # PHASE 2: Use browser automation if still no email found
        # =====================================================================
        
        if use_browser_fallback and not lead.all_emails and PLAYWRIGHT_AVAILABLE:
            logging.info(f"No email found with HTTP requests. Trying browser automation for: {website}")
            used_browser = True
            
            # URLs to try with browser
            browser_urls = [website]
            if contact_url:
                browser_urls.append(contact_url)
            
            # Try common contact URLs
            for pattern in ['/contact', '/contact-us', '/about', '/about-us']:
                test_url = urljoin(website, pattern)
                if test_url not in browser_urls:
                    browser_urls.append(test_url)
            
            # Fetch pages with browser
            browser_results = _fetch_page_with_browser_batch(browser_urls[:4], timeout=20000)
            
            for url, html in browser_results.items():
                if html:
                    _extract_from_html(html, url, lead)
                    if lead.all_emails:
                        logging.info(f"Found email using browser automation!")
                        break

    except Exception as e:
        logging.error(f"Error scraping {website}: {e}")

    # Set primary email and phone (best ones first)
    if lead.all_emails:
        lead.company_email = lead.all_emails[0]
        lead.email_status = verify_email(lead.company_email)

    if lead.all_phones:
        lead.company_phone = lead.all_phones[0]

    method = "browser" if used_browser else "HTTP"
    logging.info(f"Extracted from {lead.company_name} ({method}): {len(lead.all_emails)} emails, {len(lead.all_phones)} phones")
    return lead


def _find_additional_contact_pages(base_url: str, html: str) -> List[str]:
    """Find additional pages that might contain contact information."""
    pages = []
    soup = _make_soup(html)
    if not soup:
        return pages

    # Look for links with contact-related text
    contact_keywords = [
        'contact', 'email', 'phone', 'call', 'reach', 'support',
        'help', 'inquiry', 'enquiry', 'location', 'office', 'address',
        'get in touch', 'talk to us', 'write to us', 'connect',
    ]

    for a_tag in soup.find_all('a', href=True):
        href = a_tag.get('href', '')
        text = a_tag.get_text(strip=True).lower()
        
        # Check if link text or href contains contact keywords
        for keyword in contact_keywords:
            if keyword in text or keyword in href.lower():
                full_url = urljoin(base_url, href)
                if full_url.startswith('http') and full_url not in pages:
                    # Avoid external links
                    if _extract_domain(full_url) == _extract_domain(base_url):
                        pages.append(full_url)
                break

    return pages[:5]  # Return max 5 pages


def _fetch_page(url: str, use_browser: bool = False) -> Optional[str]:
    """Fetch a web page with retry logic. Uses browser if specified or as fallback."""
    # First try with regular requests (faster)
    if not use_browser:
        response = _make_request_with_retry(url, method='GET')
        if response and response.status_code == 200:
            # Check if we got a meaningful response (not a bot block page)
            html = response.text
            if html and len(html) > 500 and not _is_bot_blocked(html):
                return html
    
    # Fallback to browser automation for bot-protected sites
    if PLAYWRIGHT_AVAILABLE:
        logging.info(f"Using browser automation for: {url}")
        return _fetch_page_with_browser(url)
    
    return None


def _is_bot_blocked(html: str) -> bool:
    """Check if the response indicates bot detection/blocking."""
    if not html:
        return True
    
    html_lower = html.lower()
    
    # Common bot detection indicators
    bot_indicators = [
        'please enable javascript',
        'enable javascript to continue',
        'javascript is required',
        'checking your browser',
        'please wait while we verify',
        'access denied',
        'blocked',
        'captcha',
        'recaptcha',
        'cloudflare',
        'ddos protection',
        'bot detection',
        'please complete the security check',
        'unusual traffic',
        'automated access',
        'are you a robot',
        'verify you are human',
        'challenge-platform',
        'cf-browser-verification',
        'just a moment',
    ]
    
    for indicator in bot_indicators:
        if indicator in html_lower:
            logging.warning(f"Bot detection triggered: found '{indicator}'")
            return True
    
    # Check if page is too short (likely a block page)
    if len(html) < 1000 and ('javascript' in html_lower or 'redirect' in html_lower):
        return True
    
    return False


def _fetch_page_with_browser(url: str, timeout: int = 30000) -> Optional[str]:
    """Fetch page using Playwright browser automation to bypass bot detection."""
    if not PLAYWRIGHT_AVAILABLE:
        return None
    
    try:
        with sync_playwright() as p:
            # Launch browser with stealth settings
            browser = p.chromium.launch(
                headless=True,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox',
                    '--disable-setuid-sandbox',
                    '--disable-infobars',
                    '--window-size=1920,1080',
                    '--start-maximized',
                ]
            )
            
            # Create context with realistic settings
            context = browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                locale='en-US',
                timezone_id='America/New_York',
                java_script_enabled=True,
            )
            
            # Add stealth scripts to avoid detection
            context.add_init_script("""
                // Override webdriver property
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
                
                // Override plugins
                Object.defineProperty(navigator, 'plugins', {
                    get: () => [1, 2, 3, 4, 5]
                });
                
                // Override languages
                Object.defineProperty(navigator, 'languages', {
                    get: () => ['en-US', 'en']
                });
                
                // Override platform
                Object.defineProperty(navigator, 'platform', {
                    get: () => 'Win32'
                });
                
                // Override chrome property
                window.chrome = {
                    runtime: {}
                };
            """)
            
            page = context.new_page()
            
            # Navigate with wait for network idle
            try:
                page.goto(url, wait_until='networkidle', timeout=timeout)
            except PlaywrightTimeout:
                # Try with domcontentloaded if networkidle times out
                page.goto(url, wait_until='domcontentloaded', timeout=timeout)
            
            # Wait a bit for any dynamic content to load
            page.wait_for_timeout(2000)
            
            # Scroll down to trigger lazy loading
            page.evaluate("window.scrollTo(0, document.body.scrollHeight / 2)")
            page.wait_for_timeout(1000)
            
            # Get the full page content
            html = page.content()
            
            browser.close()
            
            if html and len(html) > 500:
                return html
            
    except Exception as e:
        logging.error(f"Browser automation failed for {url}: {e}")
    
    return None


def _fetch_page_with_browser_batch(urls: List[str], timeout: int = 30000) -> Dict[str, str]:
    """Fetch multiple pages using a single browser instance for efficiency."""
    results = {}
    
    if not PLAYWRIGHT_AVAILABLE or not urls:
        return results
    
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox',
                ]
            )
            
            context = browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            )
            
            # Add stealth script
            context.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                window.chrome = { runtime: {} };
            """)
            
            for url in urls:
                try:
                    page = context.new_page()
                    page.goto(url, wait_until='domcontentloaded', timeout=timeout)
                    page.wait_for_timeout(1500)
                    html = page.content()
                    if html and len(html) > 500:
                        results[url] = html
                    page.close()
                except Exception as e:
                    logging.warning(f"Failed to fetch {url}: {e}")
                    continue
            
            browser.close()
            
    except Exception as e:
        logging.error(f"Browser batch fetch failed: {e}")
    
    return results


def _extract_from_html(html: str, base_url: str, lead: CompanyLead):
    """Extract all contact information from HTML content."""
    soup = _make_soup(html)
    if not soup:
        return

    # Extract emails
    emails = extract_emails(html, soup)
    for email in emails:
        if email not in lead.all_emails:
            lead.all_emails.append(email)

    # Extract phone numbers
    phones = extract_phone_numbers(html, soup)
    for phone in phones:
        if phone not in lead.all_phones:
            lead.all_phones.append(phone)

    # Extract social links
    social_links = extract_social_links(html, soup, base_url)
    if social_links.get('linkedin') and not lead.linkedin_link:
        lead.linkedin_link = social_links['linkedin']
    if social_links.get('twitter') and not lead.twitter_link:
        lead.twitter_link = social_links['twitter']
    if social_links.get('facebook') and not lead.facebook_link:
        lead.facebook_link = social_links['facebook']
    if social_links.get('instagram') and not lead.instagram_link:
        lead.instagram_link = social_links['instagram']
    if social_links.get('tiktok') and not lead.tiktok_link:
        lead.tiktok_link = social_links['tiktok']
    if social_links.get('youtube') and not lead.youtube_link:
        lead.youtube_link = social_links['youtube']


def extract_emails(html: str, soup: BeautifulSoup) -> List[str]:
    """
    Extract email addresses from HTML content.
    Comprehensive search in mailto links, visible text, attributes, and structured data.
    """
    emails = set()

    # 1. Extract from mailto links (highest priority)
    for a_tag in soup.find_all('a', href=True):
        href = a_tag.get('href', '')
        if 'mailto:' in href.lower():
            # Handle various mailto formats
            email = href.lower().replace('mailto:', '').split('?')[0].split('#')[0].strip()
            email = re.sub(r'^/+', '', email)  # Remove leading slashes
            if _is_valid_email(email):
                emails.add(email.lower())

    # 2. Extract from visible text using regex (multiple patterns)
    text_content = soup.get_text(separator=' ', strip=True)
    
    # Standard email pattern
    for email in EMAIL_PATTERN_STRICT.findall(html):
        if _is_valid_email(email):
            emails.add(email.lower())
    
    # Also search in text content
    for email in EMAIL_PATTERN_STRICT.findall(text_content):
        if _is_valid_email(email):
            emails.add(email.lower())

    # 3. Extract from HTML attributes (data-email, data-contact, etc.)
    for tag in soup.find_all(True):
        for attr_name, attr_value in tag.attrs.items():
            if isinstance(attr_value, str):
                if 'email' in attr_name.lower() or 'mail' in attr_name.lower():
                    for email in EMAIL_PATTERN_STRICT.findall(attr_value):
                        if _is_valid_email(email):
                            emails.add(email.lower())
                # Check href and data attributes
                if attr_name in ['href', 'data-email', 'data-mail', 'data-contact', 'content']:
                    for email in EMAIL_PATTERN_STRICT.findall(attr_value):
                        if _is_valid_email(email):
                            emails.add(email.lower())

    # 4. Extract from structured data (JSON-LD, microdata)
    for script in soup.find_all('script', type='application/ld+json'):
        try:
            import json
            data = json.loads(script.string or '{}')
            _extract_email_from_json(data, emails)
        except:
            pass

    # 5. Extract from meta tags
    for meta in soup.find_all('meta'):
        content = meta.get('content', '')
        if content:
            for email in EMAIL_PATTERN_STRICT.findall(content):
                if _is_valid_email(email):
                    emails.add(email.lower())

    # 6. Search in footer specifically
    footer_selectors = [
        soup.find('footer'),
        soup.find(id=re.compile(r'footer', re.I)),
        soup.find(class_=re.compile(r'footer', re.I)),
        soup.find(class_=re.compile(r'contact', re.I)),
        soup.find(id=re.compile(r'contact', re.I)),
    ]
    for footer in footer_selectors:
        if footer:
            footer_html = str(footer)
            for email in EMAIL_PATTERN_STRICT.findall(footer_html):
                if _is_valid_email(email):
                    emails.add(email.lower())

    # 7. Search in common contact sections
    contact_selectors = [
        soup.find(class_=re.compile(r'contact-info|contact-details|contact-section', re.I)),
        soup.find(id=re.compile(r'contact-info|contact-details', re.I)),
        soup.find('address'),
    ]
    for section in contact_selectors:
        if section:
            for email in EMAIL_PATTERN_STRICT.findall(str(section)):
                if _is_valid_email(email):
                    emails.add(email.lower())

    # 8. Handle obfuscated emails (e.g., "info [at] company [dot] com")
    obfuscated = re.findall(
        r'([A-Za-z0-9._%+-]+)\s*[\[\(]?\s*at\s*[\]\)]?\s*([A-Za-z0-9.-]+)\s*[\[\(]?\s*dot\s*[\]\)]?\s*([A-Za-z]{2,})',
        html, re.I
    )
    for parts in obfuscated:
        email = f"{parts[0]}@{parts[1]}.{parts[2]}".lower()
        if _is_valid_email(email):
            emails.add(email)

    # 9. Handle JavaScript-protected emails (common pattern)
    js_emails = re.findall(
        r"['\"]([A-Za-z0-9._%+-]+)['\"].*?['\"]@['\"].*?['\"]([A-Za-z0-9.-]+\.[A-Za-z]{2,})['\"]",
        html
    )
    for parts in js_emails:
        email = f"{parts[0]}@{parts[1]}".lower()
        if _is_valid_email(email):
            emails.add(email)

    # Filter out common non-business emails
    filtered = [e for e in emails if not _is_generic_email(e)]
    
    # Sort by priority (info@, contact@, hello@ first)
    priority_prefixes = ['info@', 'contact@', 'hello@', 'sales@', 'support@', 'enquiry@', 'inquiry@']
    sorted_emails = sorted(filtered, key=lambda e: (
        0 if any(e.startswith(p) for p in priority_prefixes) else 1,
        e
    ))
    
    return sorted_emails[:10]  # Return up to 10 emails


def extract_phone_numbers(html: str, soup: BeautifulSoup) -> List[str]:
    """
    Extract phone numbers from HTML content.
    Comprehensive search in tel links, visible text, attributes, and structured data.
    """
    phones = set()

    # 1. Extract from tel links (highest priority)
    for a_tag in soup.find_all('a', href=True):
        href = a_tag.get('href', '')
        if 'tel:' in href.lower():
            phone = href.lower().replace('tel:', '').replace('%20', '').strip()
            phone = re.sub(r'[^\d+]', '', phone)  # Keep only digits and +
            if len(phone) >= 7:
                phones.add(_format_phone(phone))
            
            # Also check the link text
            link_text = a_tag.get_text(strip=True)
            cleaned = _clean_phone_number(link_text)
            if cleaned:
                phones.add(cleaned)

    # 2. Extract from HTML attributes
    for tag in soup.find_all(True):
        for attr_name, attr_value in tag.attrs.items():
            if isinstance(attr_value, str):
                if 'phone' in attr_name.lower() or 'tel' in attr_name.lower():
                    cleaned = _clean_phone_number(attr_value)
                    if cleaned:
                        phones.add(cleaned)
                if attr_name in ['data-phone', 'data-tel', 'data-number', 'content']:
                    for pattern in PHONE_PATTERNS:
                        for match in pattern.findall(attr_value):
                            cleaned = _clean_phone_number(match)
                            if cleaned:
                                phones.add(cleaned)

    # 3. Extract from visible text using multiple regex patterns
    text_content = soup.get_text(separator=' ', strip=True)
    
    for pattern in PHONE_PATTERNS:
        # Search in raw HTML
        for match in pattern.findall(html):
            cleaned = _clean_phone_number(match)
            if cleaned:
                phones.add(cleaned)
        # Search in text content
        for match in pattern.findall(text_content):
            cleaned = _clean_phone_number(match)
            if cleaned:
                phones.add(cleaned)

    # 4. Extract from structured data (JSON-LD)
    for script in soup.find_all('script', type='application/ld+json'):
        try:
            import json
            data = json.loads(script.string or '{}')
            _extract_phone_from_json(data, phones)
        except:
            pass

    # 5. Extract from meta tags
    for meta in soup.find_all('meta'):
        content = meta.get('content', '')
        if content:
            for pattern in PHONE_PATTERNS:
                for match in pattern.findall(content):
                    cleaned = _clean_phone_number(match)
                    if cleaned:
                        phones.add(cleaned)

    # 6. Search in footer and contact sections
    selectors = [
        soup.find('footer'),
        soup.find(id=re.compile(r'footer', re.I)),
        soup.find(class_=re.compile(r'footer', re.I)),
        soup.find(class_=re.compile(r'contact', re.I)),
        soup.find(id=re.compile(r'contact', re.I)),
        soup.find('address'),
        soup.find(class_=re.compile(r'phone|tel|call', re.I)),
    ]
    
    for section in selectors:
        if section:
            section_html = str(section)
            for pattern in PHONE_PATTERNS:
                for match in pattern.findall(section_html):
                    cleaned = _clean_phone_number(match)
                    if cleaned:
                        phones.add(cleaned)

    # 7. Look for phone labels and nearby text
    phone_labels = soup.find_all(string=re.compile(r'phone|tel|call|mobile|fax|contact', re.I))
    for label in phone_labels:
        parent = label.parent
        if parent:
            # Check siblings and parent text
            context = str(parent.parent) if parent.parent else str(parent)
            for pattern in PHONE_PATTERNS:
                for match in pattern.findall(context):
                    cleaned = _clean_phone_number(match)
                    if cleaned:
                        phones.add(cleaned)

    # Filter and sort phones
    valid_phones = [p for p in phones if _is_valid_phone(p)]
    
    return valid_phones[:5]  # Return up to 5 phone numbers


def _format_phone(phone: str) -> str:
    """Format phone number consistently."""
    digits = re.sub(r'[^\d+]', '', phone)
    if digits.startswith('+'):
        return digits
    return digits


def _is_valid_phone(phone: str) -> bool:
    """Check if phone number is valid."""
    digits = re.sub(r'[^\d]', '', phone)
    # Must have 7-15 digits
    if len(digits) < 7 or len(digits) > 15:
        return False
    # Avoid false positives (years, zip codes, etc.)
    if re.match(r'^(19|20)\d{2}$', digits):  # Years
        return False
    if len(digits) == 5 and digits.isdigit():  # Zip codes
        return False
    return True


def extract_social_links(html: str, soup: BeautifulSoup, base_url: str) -> Dict[str, str]:
    """
    Extract social media links from HTML content.
    Searches anchor tags, footer sections, and visible URLs.
    """
    social_links = {
        'linkedin': '',
        'twitter': '',
        'facebook': '',
        'instagram': '',
        'tiktok': '',
        'youtube': '',
    }

    # Collect all href values
    all_hrefs = []
    for a_tag in soup.find_all('a', href=True):
        href = a_tag.get('href', '').strip()
        if href.startswith('http'):
            all_hrefs.append(href)
        elif href.startswith('/'):
            all_hrefs.append(urljoin(base_url, href))

    href_text = ' '.join(all_hrefs)

    # Match social patterns
    for platform, pattern in SOCIAL_PATTERNS.items():
        match = pattern.search(href_text)
        if match:
            url = match.group(0).strip()
            # Clean trailing characters
            url = re.sub(r'[)\]\'",;]+$', '', url)
            social_links[platform] = url

    # Also search in raw HTML for social links
    for platform, pattern in SOCIAL_PATTERNS.items():
        if not social_links[platform]:
            match = pattern.search(html)
            if match:
                url = match.group(0).strip()
                url = re.sub(r'[)\]\'",;]+$', '', url)
                social_links[platform] = url

    return social_links


# =============================================================================
# EMAIL VERIFICATION
# =============================================================================

def verify_email(email: str) -> str:
    """
    Verify email address using two-level validation:
    Level 1: Regex validation
    Level 2: MX record check
    Returns: "Verified" or "Not verified"
    """
    if not email:
        return "Not verified"

    # Level 1: Regex validation
    if not _is_valid_email(email):
        return "Not verified"

    # Level 2: MX record check
    try:
        domain = email.split('@')[1]
        dns.resolver.resolve(domain, 'MX')
        return "Verified"
    except dns.resolver.NXDOMAIN:
        return "Not verified"
    except dns.resolver.NoAnswer:
        return "Not verified"
    except dns.resolver.NoNameservers:
        return "Not verified"
    except Exception as e:
        logging.debug(f"MX check failed for {email}: {e}")
        return "Not verified"


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def save_to_excel(leads: List[CompanyLead]) -> tuple:
    """
    Save leads to professionally styled Excel file named Scraping.xlsx.
    Returns (added_count, error_message).
    """
    try:
        # Load existing or create new workbook
        if os.path.exists(MASTER_DB_FILE):
            wb = load_workbook(MASTER_DB_FILE)
            ws = wb.active
            existing_websites = _get_existing_websites(ws)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            existing_websites = set()

            # Add headers with styling
            _setup_header_row(ws)

        # Add new leads (avoid duplicates)
        added_count = 0
        for lead in leads:
            website_key = lead.company_website.strip().lower()
            if website_key and website_key not in existing_websites:
                row_num = ws.max_row + 1
                row_data = lead.to_excel_row()

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=_sanitize_cell(value))

                    # Apply borders
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # Color code email status (column 4)
                    if col_idx == 4:
                        if value == "Verified":
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

                existing_websites.add(website_key)
                added_count += 1

        # Apply final formatting
        _apply_excel_styling(ws)

        # Save file
        wb.save(MASTER_DB_FILE)
        logging.info(f"Saved {added_count} leads to {MASTER_DB_FILE}")

        return added_count, None

    except Exception as e:
        logging.error(f"Error saving to Excel: {e}")
        return 0, str(e)


def _setup_header_row(ws):
    """Setup header row with professional styling."""
    # Header styling
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = 'A2'


def _apply_excel_styling(ws):
    """Apply final styling to the worksheet."""
    # Auto-adjust column widths
    column_widths = {
        1: 30,   # Company Name
        2: 40,   # Company website
        3: 35,   # Company Email
        4: 15,   # Email status
        5: 20,   # Company Phone Number
        6: 40,   # Linkedin Link
        7: 35,   # X/ Twitter Link
        8: 35,   # Facebook Link
        9: 35,   # Instagram Link
        10: 35,  # Tiktok Link
        11: 40,  # YouTube channel Link
    }

    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _get_existing_websites(ws) -> Set[str]:
    """Get set of existing website URLs from worksheet."""
    existing = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]
        if cell.value:
            existing.add(str(cell.value).strip().lower())
    return existing


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def _make_soup(html_content: str) -> Optional[BeautifulSoup]:
    """Create BeautifulSoup object from HTML."""
    if not html_content:
        return None
    try:
        return BeautifulSoup(html_content, "lxml")
    except:
        return BeautifulSoup(html_content, "html.parser")


def _make_request_with_retry(url: str, params: Dict = None, method: str = 'GET') -> Optional[requests.Response]:
    """Make HTTP request with retry logic and timeout handling."""
    for attempt in range(MAX_RETRIES):
        try:
            if method == 'GET':
                response = requests.get(
                    url,
                    params=params,
                    headers=get_headers(),
                    timeout=REQUEST_TIMEOUT,
                    allow_redirects=True
                )
            else:
                response = requests.post(
                    url,
                    json=params,
                    headers=get_headers(),
                    timeout=REQUEST_TIMEOUT
                )

            if response.status_code == 200:
                return response
            elif response.status_code == 429:  # Rate limited
                time.sleep(RETRY_DELAY * (attempt + 1))
            else:
                return response

        except requests.exceptions.Timeout:
            logging.warning(f"Timeout on attempt {attempt + 1} for {url}")
            time.sleep(RETRY_DELAY)
        except requests.exceptions.RequestException as e:
            logging.warning(f"Request error on attempt {attempt + 1} for {url}: {e}")
            time.sleep(RETRY_DELAY)

    return None


def _extract_domain(url: str) -> str:
    """Extract domain from URL."""
    try:
        parsed = urlparse(url)
        domain = parsed.netloc or parsed.path.split('/')[0]
        return domain.replace('www.', '').lower()
    except:
        return ''


def _is_valid_company_website(url: str) -> bool:
    """Check if URL looks like a valid company website."""
    if not url:
        return False

    # Exclude social media, aggregator sites, etc.
    excluded_domains = [
        'facebook.com', 'twitter.com', 'x.com', 'linkedin.com',
        'instagram.com', 'youtube.com', 'tiktok.com', 'pinterest.com',
        'yelp.com', 'yellowpages.com', 'wikipedia.org', 'amazon.com',
        'ebay.com', 'craigslist.org', 'reddit.com', 'quora.com',
    ]

    domain = _extract_domain(url)
    for excluded in excluded_domains:
        if excluded in domain:
            return False

    return True


def _is_valid_email(email: str) -> bool:
    """Validate email format."""
    if not email or '@' not in email:
        return False

    email = email.strip().lower()
    
    # Check for image/file extensions
    invalid_extensions = (
        '.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg', '.ico',
        '.pdf', '.doc', '.docx', '.css', '.js', '.php', '.html'
    )
    if email.endswith(invalid_extensions):
        return False
    
    # Check for common false positives
    if email.startswith(('http://', 'https://', 'www.', '//')):
        return False
    
    # Must have valid structure
    parts = email.split('@')
    if len(parts) != 2:
        return False
    
    local_part, domain = parts
    
    # Local part validation
    if not local_part or len(local_part) > 64:
        return False
    
    # Domain validation
    if not domain or '.' not in domain:
        return False
    
    domain_parts = domain.split('.')
    if len(domain_parts[-1]) < 2:  # TLD must be at least 2 chars
        return False
    
    # Check against invalid domains
    invalid_domains = [
        'example.com', 'test.com', 'localhost', 'domain.com',
        'email.com', 'mail.com', 'yoursite.com', 'yourdomain.com',
        'sentry.io', 'wixpress.com', 'sentry-next.wixpress.com',
    ]
    if domain in invalid_domains:
        return False

    # Basic regex check
    if not EMAIL_PATTERN_STRICT.match(email):
        return False

    return True


def _is_generic_email(email: str) -> bool:
    """Check if email is a generic/non-business address that should be filtered out."""
    email_lower = email.lower()
    
    # Prefixes that indicate non-useful emails
    filter_prefixes = [
        'noreply@', 'no-reply@', 'donotreply@', 'do-not-reply@',
        'mailer-daemon@', 'postmaster@', 'hostmaster@',
        'abuse@', 'spam@', 'unsubscribe@',
        'example@', 'test@', 'demo@', 'sample@',
        'user@', 'username@', 'email@', 'mail@',
        'placeholder@', 'dummy@', 'fake@',
    ]
    
    # Domains that indicate non-useful emails
    filter_domains = [
        '@example.com', '@test.com', '@localhost', '@domain.com',
        '@email.com', '@mail.com', '@yoursite.com', '@yourdomain.com',
        '@sentry.io', '@wixpress.com', '@placeholder.com',
        '@mailinator.com', '@tempmail.com', '@guerrillamail.com',
    ]
    
    # Check prefixes
    for prefix in filter_prefixes:
        if email_lower.startswith(prefix):
            return True
    
    # Check domains
    for domain in filter_domains:
        if domain in email_lower:
            return True
    
    return False


def _clean_phone_number(phone: str) -> str:
    """Clean and normalize phone number."""
    if not phone:
        return ''

    # Remove common non-digit characters but keep the original format for display
    original = phone.strip()
    
    # Extract digits only for validation
    digits_only = re.sub(r'[^\d]', '', phone)
    
    # Must have at least 7 digits and no more than 15
    if len(digits_only) < 7 or len(digits_only) > 15:
        return ''
    
    # Avoid false positives (years, zip codes, etc.)
    if re.match(r'^(19|20)\d{2}$', digits_only):  # Years like 2024
        return ''
    if len(digits_only) == 5:  # Zip codes
        return ''
    
    # Return cleaned but formatted version
    # Keep + prefix if present, otherwise just return digits with formatting
    if original.startswith('+'):
        return '+' + digits_only
    
    # Format US numbers nicely
    if len(digits_only) == 10:
        return f"({digits_only[:3]}) {digits_only[3:6]}-{digits_only[6:]}"
    elif len(digits_only) == 11 and digits_only.startswith('1'):
        return f"+1 ({digits_only[1:4]}) {digits_only[4:7]}-{digits_only[7:]}"
    
    return original.strip()


def _find_contact_page(base_url: str, html: str) -> Optional[str]:
    """Find contact page URL from homepage."""
    soup = _make_soup(html)
    if not soup:
        return None

    # Search for contact links
    for a_tag in soup.find_all('a', href=True):
        href = a_tag.get('href', '').lower()
        text = a_tag.get_text().lower()

        for pattern in CONTACT_PAGE_PATTERNS:
            if pattern in href or 'contact' in text:
                full_url = urljoin(base_url, a_tag.get('href', ''))
                if full_url.startswith('http'):
                    return full_url

    # Try common contact page URLs
    for pattern in CONTACT_PAGE_PATTERNS:
        test_url = urljoin(base_url, pattern)
        try:
            response = requests.head(test_url, headers=get_headers(), timeout=5, allow_redirects=True)
            if response.status_code == 200:
                return test_url
        except:
            pass

    return None


def _find_about_page(base_url: str, html: str) -> Optional[str]:
    """Find about page URL from homepage."""
    soup = _make_soup(html)
    if not soup:
        return None

    # Search for about links
    for a_tag in soup.find_all('a', href=True):
        href = a_tag.get('href', '').lower()
        text = a_tag.get_text().lower()

        for pattern in ABOUT_PAGE_PATTERNS:
            if pattern in href or 'about' in text:
                full_url = urljoin(base_url, a_tag.get('href', ''))
                if full_url.startswith('http'):
                    return full_url

    return None


def _extract_email_from_json(data, emails_set: Set):
    """Recursively extract emails from JSON-LD data."""
    if isinstance(data, dict):
        for key, value in data.items():
            if key.lower() == 'email' and isinstance(value, str):
                if _is_valid_email(value):
                    emails_set.add(value.lower())
            else:
                _extract_email_from_json(value, emails_set)
    elif isinstance(data, list):
        for item in data:
            _extract_email_from_json(item, emails_set)


def _extract_phone_from_json(data, phones_set: Set):
    """Recursively extract phones from JSON-LD data."""
    if isinstance(data, dict):
        for key, value in data.items():
            if key.lower() in ('telephone', 'phone', 'phonenumber') and isinstance(value, str):
                cleaned = _clean_phone_number(value)
                if cleaned:
                    phones_set.add(cleaned)
            else:
                _extract_phone_from_json(value, phones_set)
    elif isinstance(data, list):
        for item in data:
            _extract_phone_from_json(item, phones_set)


def _sanitize_cell(value) -> str:
    """Sanitize cell value for Excel."""
    if value is None:
        return ""
    s = str(value).strip()
    s = s.replace("\x00", "").replace("\r\n", "\n").replace("\r", "\n")
    s = "".join(c for c in s if c == "\n" or (ord(c) >= 32 and ord(c) != 127))
    return s[:32767]  # Excel cell limit


# =============================================================================
# MAIN SCRAPING WORKFLOW
# =============================================================================

def scrape_leads(query: str, limit: int = 10) -> List[CompanyLead]:
    """
    Main function to scrape leads for a given query.
    1. Discover companies via Google API
    2. Extract contact info via HTML parsing
    3. Verify emails
    4. Return list of CompanyLead objects
    """
    logging.info(f"Starting lead scrape for query: {query}")

    # Step 1: Discover companies using Google API
    companies = search_companies_google(query, limit * 2)  # Get extra for filtering
    logging.info(f"Discovered {len(companies)} companies from Google API")

    # Step 2: Filter and validate websites
    valid_companies = get_company_websites(companies)
    logging.info(f"Validated {len(valid_companies)} company websites")

    # Step 3: Scrape contact info from each website (parallel)
    leads = []
    seen_domains = set()

    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        future_to_company = {
            executor.submit(scrape_website_contact_info, company): company
            for company in valid_companies[:limit]
        }

        for future in concurrent.futures.as_completed(future_to_company):
            try:
                lead = future.result()
                domain = _extract_domain(lead.company_website)

                # Avoid duplicates
                if domain and domain not in seen_domains:
                    seen_domains.add(domain)
                    leads.append(lead)
            except Exception as e:
                logging.error(f"Error processing lead: {e}")

    logging.info(f"Scraped contact info for {len(leads)} leads")
    return leads


# =============================================================================
# FLASK API ENDPOINTS
# =============================================================================

@app.route('/api/search-leads', methods=['POST'])
def search_leads_api():
    """
    API endpoint to search and scrape leads.
    Request body: {"query": "software houses in NYC", "limit": 10}
    """
    data = request.get_json() or {}
    query = data.get('query')
    limit = int(data.get('limit', 10))

    if not query:
        return jsonify({"error": "Query is required"}), 400

    start_time = time.time()

    try:
        # Scrape leads
        leads = scrape_leads(query, limit)

        # Save to Excel
        added_count, save_error = save_to_excel(leads)

        # Prepare response
        response = {
            "status": "success",
            "duration_seconds": round(time.time() - start_time, 2),
            "total_found": len(leads),
            "new_leads_added": added_count,
            "file_path": os.path.abspath(MASTER_DB_FILE),
            "leads": [lead.to_dict() for lead in leads],
        }

        if save_error:
            response["save_error"] = save_error

        return jsonify(response)

    except Exception as e:
        logging.exception("Search leads error")
        return jsonify({
            "status": "error",
            "error": str(e),
            "duration_seconds": round(time.time() - start_time, 2),
        }), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({
        "status": "healthy",
        "google_api_configured": bool(GOOGLE_API_KEY),
        "google_cse_configured": bool(GOOGLE_CSE_ID),
    })


# =============================================================================
# CLI INTERFACE
# =============================================================================

def main():
    """Command-line interface for the scraper."""
    import argparse

    parser = argparse.ArgumentParser(description='Lead Scraping System')
    parser.add_argument('--query', '-q', type=str, help='Search query (e.g., "software houses in NYC")')
    parser.add_argument('--limit', '-l', type=int, default=10, help='Number of leads to scrape')
    parser.add_argument('--server', '-s', action='store_true', help='Run as Flask server')
    parser.add_argument('--port', '-p', type=int, default=5000, help='Server port')

    args = parser.parse_args()

    if args.server:
        logging.info(f"Starting server on port {args.port}")
        app.run(debug=False, port=args.port, host='0.0.0.0')
    elif args.query:
        logging.info(f"Scraping leads for: {args.query}")
        leads = scrape_leads(args.query, args.limit)

        if leads:
            added_count, error = save_to_excel(leads)
            print(f"\n{'='*60}")
            print(f"Scraping Complete!")
            print(f"{'='*60}")
            print(f"Query: {args.query}")
            print(f"Total leads found: {len(leads)}")
            print(f"New leads added to Excel: {added_count}")
            print(f"Output file: {os.path.abspath(MASTER_DB_FILE)}")

            if error:
                print(f"Save error: {error}")

            print(f"\n{'='*60}")
            print("Sample leads:")
            print(f"{'='*60}")
            for i, lead in enumerate(leads[:5], 1):
                print(f"\n{i}. {lead.company_name}")
                print(f"   Website: {lead.company_website}")
                print(f"   Email: {lead.company_email} ({lead.email_status})")
                print(f"   Phone: {lead.company_phone}")
        else:
            print("No leads found.")
    else:
        parser.print_help()

@app.route('/scrape', methods=['POST'])
def scrape_api():
    data = request.get_json()
    query = data.get('query')
    limit = data.get('limit', 10)

    if not query:
        return jsonify({"error": "Query is required"}), 400

    # Yahan aapka scraping logic call hoga
    results = search_companies_google(query, limit)
    # Baqi logic...
    
    return jsonify({"status": "success", "message": f"Scraping started for {query}"})

if __name__ == '__main__':
    app.run(debug=True, port=5000)


if __name__ == '__main__':
    main()
